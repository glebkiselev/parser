#!/usr/bin/python
from urllib.request import urlopen
from urllib import error
import re

import itertools
from bs4 import BeautifulSoup
import csv

from flask import Flask
from flask import request, render_template, redirect, url_for, jsonify
from openpyxl import load_workbook
import pandas
from multiprocessing import Pool, Process
from multiprocessing.dummy import Pool as ThreadPool
import subprocess


UPLOAD_FOLDER = 'data'
ALLOWED_EXTENSIONS = set(['txt', 'pdf', 'csv', 'xls', 'xlsx', 'docx'])

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/app.js')
def sencha_app():
    return redirect(url_for('static', filename='app.js'))


def parse_links(links, visited, url):
    # убираем ссылки на на js и  css файлы
    timelist = []
    truelinks = []
    copy_links = links.copy()
    for link in copy_links:
        timelist.extend(link.split("/"))
        if ".js" in timelist[len(timelist) - 1].lower() or ".css" in timelist[len(timelist) - 1].lower():
            links.remove(link)
        elif "." in timelist[len(timelist) - 1].lower():
            timelist = timelist[len(timelist) - 1].lower().split(".")
            if "html" not in timelist and "php" not in timelist:
                links.remove(link)
        elif link in visited:
            links.remove(link)
        elif len(re.findall("[.,\-\s\?*\{\}\#]", timelist[len(timelist) - 1].lower())):
            links.remove(link)
        else:
            truelinks.append(link)
    # чистит список от ссылок на левые сайты (почти все)
    copy = truelinks.copy()
    for link in copy:
        if link.startswith("http://") or link.startswith("https://"):
            if not url.split("/")[2] == link.split("/")[2]:
                truelinks.remove(link)
            # и от самого себя
            elif url == link:
                truelinks.remove(link)
            elif not link.startswith(url):
                truelinks.remove(link)
        elif not link.startswith("//ww"):
            truelinks.append(url + link[1:])
            truelinks.remove(link)
        else:
            truelinks.remove(link)
    return truelinks


def get_links(html, url, visited):
    links = [link[0] for link in list(set(re.findall('"((http|ftp)s?://.*?)"', html)))]
    nonfull = list(set(re.findall('href="(.*?)"', html)))

    truelinks = parse_links(links, visited, url)
    truelinks2 = parse_links(nonfull, visited, url)

    truelinks2.extend(truelinks)
    truelinks2 = list(set(truelinks2))

    return truelinks2


def get_text(html):
    if html:
        soup = BeautifulSoup(html, "lxml")

        # kill all script and style elements
        for script in soup(["script", "style"]):
            script.extract()  # rip it out

        # get all div
        div_saver = []
        for div in soup(["div"]):
            div = div.get_text()
            div_saver.append(div)



        div_saver2 = div_saver.copy()

        for text in div_saver2:
            # get text - если soup.body.get_text() - вернет только боди тд по аналогии
            # text = soup.get_text()

            # break into lines and remove leading and trailing space on each
            lines = (line.strip() for line in text.splitlines())
            # break multi-headlines into a line each
            chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
            # drop blank lines
            text2 = ".".join(chunk for chunk in chunks if chunk)
            div_saver.remove(text)
            div_saver.append(text2)

        return div_saver
    else:
        return " "

def read_url(url):
    try:
        with urlopen(url) as data:
            enc = data.info().get_content_charset('utf-8')
            try:
                html = data.read().decode(enc)
            except UnicodeDecodeError:
                print("bad unicode decoding")
                return None
    except UnicodeEncodeError:
        print("bad unicode decoding")
        return None
    return html


def find_words(url, text, words, posts):
    for word in words:
        if word in text:
            if url not in posts:
                pattern = re.compile(word, re.IGNORECASE)
                text = pattern.sub(word, text)
                posts.append((url, text))
    return posts


def main_alg(url, link, words, posts, visited_links, depth):
    if depth == 0:
        return posts
    try:
        # читаем ссылку и получаем ее html код
        html = read_url(link)
    except error.URLError as err:
        return posts
    except UnicodeEncodeError:
        return posts
    if not html:
        html = ''
    # получаем строку с текстом в cсылке
    text_full = list(set(get_text(html)))
    # ищем слова и получаем назад список постов, где они были найдены
    for text in text_full:
        fwords = find_words(link, text, words, posts)
        if fwords and fwords[0] not in posts:
            posts.extend([list(set(fwords) - set(posts)), text])
        posts = list(set(posts))
        print("posts are found: {0}".format(len(posts)))

        # for post in posts:
        # print(post)
    # получаем все ссылки ресурса
    links = get_links(html, url, visited_links)

    links2 = []
    count = 0
    if len(links) > 10:
        while count < 10:
            links2.append(links[count])
            count += 1
    else:
        links2 = links
    for link in links2:
        try:
            visited_links.append(link)
            posts.extend(main_alg(url, link, words, posts, visited_links, depth - 1))
        except error.HTTPError as e:
            print("bad request")
    return posts


@app.route('/_findwords')
def add_numbers():
    #get links and keywords
    a= execsear()
    urls=[]
    urls.extend(a[0])
    words = []
    words.extend(a[1])

    depth = 2  # размерность поиска вглубину (кол-во страниц сайта, которые мы просмотрим)
    # read from csv marks and models
    cars_file = "static/cars_csv.csv"
    marks_and_models = read_csv(cars_file)


    if len(urls) > 1:
            # разбиваю по 10 итого 50 url
            urls_for_process = list(group(urls, 5))
            iterater = 0
            for url in urls_for_process:
                # for xlsx file:
                #filename = 'reader.xlsx'
                # for csv files
                filename = str(iterater) + ".csv"
                # for xlsx file: filename
                subprocess.call(['touch', filename])
                post_searcher(url, words, depth, marks_and_models, filename)
                iterater+=1






def group(iterable, count):
    return zip(*[iter(iterable)] * count)




def post_searcher(urls, words, depth, marks_and_models, filename):
    for url in urls:
        posts = []
        visited_links = [url]  # was here
        if url[len(url) - 1] == " ":
            url = del_end_probel(url)
        if url[0] == " ":
            url = del_start_probel(url)
        if not url.startswith("http://") and not url.startswith("https://"):
            url = 'http://' + url

        posts = set(main_alg(url, url, words, posts, visited_links, depth))
        if len(posts) < 10:
            continue
        print("url:{0}  for post len: {1}".format(url, len(posts)))
        good = []
        costs = ["cost", "цен", "скид", "руб", "процент", "клиент", "%"]
        compons = list(itertools.product(marks_and_models, costs))
        posts = list(posts)
        for post in posts:
            strings = post[1].split(".")
            for string in strings:
                for compon in compons:
                    if compon[0][0] in string and compon[0][1] in post[1] and compon[1] in post[1]:
                        good.append((post[0], compon[0][0], compon[0][1], post[1]))

        # to write for excisting xlsx
        #write_xlsx(good, filename)
        # to wite for excisting csv
        write_csv(good, filename)

def write_csv(good, filename):
    with open(filename, 'a') as resultFile:
        wr = csv.writer(resultFile, dialect='excel')
        wr.writerow(good)


def write_xlsx(good, filename):
    labels = ['link', 'mark', 'model', 'post']
    df = pandas.DataFrame.from_records(good, columns=labels)
    book = load_workbook(filename)
    writer = pandas.ExcelWriter(filename, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    df.to_excel(writer, "Main", columns=labels)

    writer.save()


def read_csv(cars_file):
    with open(cars_file, encoding='mac_roman') as f:
        datas = [line for line in csv.reader(f)]
    list_data = []
    for data in datas:
        data = data[0].split(";")
        if data[0]:
            list_data.append((data[0], data[1]))

    return list_data


def del_start_probel(url):
    url = url[1:]
    if url[0] == " ":
       url = del_start_probel(url)
       return url
    else:
        return url

def del_end_probel(url):
    url = url[:-1]
    if url[len(url) - 1] == " ":
       url = del_end_probel(url)
       return url
    else:
        return url

@app.route('/')
def index():
    return render_template('index.html')


def execsear():
    wb = load_workbook('static/test_check.xlsx')
    ws = wb['data']
    links = []
    words = []
    for col in ws['A']:
        if not col.value == None:
            links.append(col.value)
    for col in ws['B']:
        if not col.value == None:
            words.append(col.value)

    return [links, words]



if __name__ == '__main__':
    # включает веб морду
    app.run()
    # для тестов без веб-морды
    # run('http://toyota-axsel.com', ["price"])
